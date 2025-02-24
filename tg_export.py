"""
====================================
    è¶‹äºé£æœºæ•°æ®ä¸€é”®å¯¼å‡ºå·¥å…· v1.0.0   
====================================

âœ¨ åŠŸèƒ½ç®€ä»‹
----------
ä¸€æ¬¾åŠŸèƒ½å¼ºå¤§çš„ Telegram æ•°æ®å¯¼å‡ºå·¥å…·ï¼Œæ”¯æŒå¤šç§æ•°æ®ç±»å‹å¯¼å‡ºå’Œ Excel æ ¼å¼ä¿å­˜ã€‚

ğŸ“± Telegramç¤¾ç¾¤ä¿¡æ¯
----------
â€¢ å®˜æ–¹é¢‘é“ï¼š@QUYUkjpd
â€¢ äº¤æµç¾¤ç»„ï¼š@QUYUkjq
â€¢ è”ç³»ä½œè€…ï¼š@Lawofforce

ğŸ’ èµåŠ©æ”¯æŒ
----------
æ„Ÿè°¢æ‚¨çš„æ”¯æŒï¼Œè¿™æ˜¯æˆ‘ä»¬æŒç»­æ”¹è¿›çš„åŠ¨åŠ›ï¼

â€¢ TRC20-USDT é’±åŒ…åœ°å€:
  TQ2gs6167orQSVWVNHWrKq9SZ8a5WRETZs

âš ï¸ å…è´£å£°æ˜
----------
â€¢ æœ¬å·¥å…·ä»…ä¾›å­¦ä¹ äº¤æµä½¿ç”¨
â€¢ ä¸¥ç¦ç”¨äºéæ³•ç”¨é€”
â€¢ ä½¿ç”¨æœ¬å·¥å…·æ‰€äº§ç”Ÿçš„ä¸€åˆ‡åæœç”±ä½¿ç”¨è€…è‡ªè¡Œæ‰¿æ‹…

ğŸ“œ è®¸å¯åè®®
----------
MIT License

Copyright (c) 2025 git88848

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
"""

import sys
sys.dont_write_bytecode = True  # ç¦æ­¢ç”Ÿæˆ __pycache__

from telethon.tl.functions.messages import GetFullChatRequest
from telethon.tl.functions.channels import GetFullChannelRequest
from telethon.sync import TelegramClient
from telethon.tl.types import User, Channel, Chat
import openpyxl
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

# Telegram API é…ç½®
# è¯·ä» https://my.telegram.org/apps è·å–è¿™äº›å€¼
API_ID = 28338165  # ä¿®æ”¹ä¸ºæ–°çš„ API_ID
API_HASH = 'f26de618b5433c51826eb4123400ddf'  # ä¿®æ”¹ä¸ºæ–°çš„ API_HASH

def adjust_column_width(worksheet):
    """è‡ªåŠ¨è°ƒæ•´åˆ—å®½"""
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width

async def get_chat_invite_link(client, entity):
    """è·å–ç¾¤ç»„/é¢‘é“çš„é‚€è¯·é“¾æ¥"""
    try:
        # é¦–å…ˆå°è¯•è·å–å…¬å¼€é“¾æ¥
        if hasattr(entity, 'username') and entity.username:
            return f"https://t.me/{entity.username}"
        
        # å¦‚æœæ˜¯è¶…çº§ç¾¤ç»„æˆ–é¢‘é“ï¼Œå°è¯•è·å–é‚€è¯·é“¾æ¥
        if isinstance(entity, Channel):
            try:
                full_entity = await client(GetFullChannelRequest(entity))
                if hasattr(full_entity.full_chat, 'invite_link') and full_entity.full_chat.invite_link:
                    return full_entity.full_chat.invite_link
            except Exception as e:
                print(f"è·å–é‚€è¯·é“¾æ¥æ—¶å‡ºé”™: {str(e)}")
        
        # å¦‚æœæ˜¯æ™®é€šç¾¤ç»„ï¼Œå°è¯•è·å–é‚€è¯·é“¾æ¥
        elif isinstance(entity, Chat):
            try:
                full_chat = await client(GetFullChatRequest(entity.id))
                if hasattr(full_chat.full_chat, 'invite_link') and full_chat.full_chat.invite_link:
                    return full_chat.full_chat.invite_link
            except Exception as e:
                print(f"è·å–ç¾¤ç»„é“¾æ¥æ—¶å‡ºé”™: {str(e)}")
    except Exception as e:
        print(f"è·å–é“¾æ¥æ—¶å‡ºé”™: {str(e)}")
    return 'æ— é“¾æ¥'

async def export_data(
    export_contacts=True,
    export_groups=True,
    export_channels=True,
    export_bots=True,
    stop_check=None
):
    print("æ­£åœ¨å¯åŠ¨ Telegram å®¢æˆ·ç«¯...")
    client = TelegramClient('session_name', API_ID, API_HASH)
    
    # æ·»åŠ ä¸­æ–‡æç¤º
    client.flood_sleep_threshold = 0
    client.parse_mode = 'html'
    
    # é‡å†™ TelegramClient çš„ _parse_phone_number æ–¹æ³•æ¥ä¿®æ”¹æç¤º
    def custom_parse_phone(self, phone):
        return phone
    
    # é‡å†™ TelegramClient çš„ start æ–¹æ³•æ¥ä¿®æ”¹æç¤º
    original_start = client.start
    async def custom_start():
        try:
            while True:
                try:
                    return await original_start(
                        phone=lambda: input("è¯·è¾“å…¥æ‰‹æœºå· (æ ¼å¼å¦‚: +8613812345678): ").strip(),
                        password=lambda: input("è¯·è¾“å…¥ä¸¤æ­¥éªŒè¯å¯†ç : "),
                        code_callback=lambda: input("è¯·è¾“å…¥æ”¶åˆ°çš„éªŒè¯ç : ").strip(),
                        first_name=lambda: "User",
                        last_name=lambda: "Name"
                    )
                except EOFError:
                    print("\né”™è¯¯ï¼šè¾“å…¥è¢«ä¸­æ–­")
                    print("è¯·é‡æ–°è¾“å…¥æ‰‹æœºå·")
                    continue
        except Exception as e:
            error_msg = str(e).lower()
            if "phone number has been banned" in error_msg:
                print("\né”™è¯¯ï¼šæ­¤æ‰‹æœºå·å·²è¢« Telegram å°ç¦ï¼Œæ— æ³•ç»§ç»­ä½¿ç”¨ã€‚")
                print("å»ºè®®ï¼š")
                print("1. ä½¿ç”¨å…¶ä»–æœªè¢«å°ç¦çš„æ‰‹æœºå·")
                print("2. å¦‚æœè®¤ä¸ºæ˜¯è¯¯å°ï¼Œå¯ä»¥è®¿é—® https://www.telegram.org/faq_spam ç”³è¯‰")
                input("\næŒ‰å›è½¦é”®é€€å‡º...")
                sys.exit(1)
            else:
                print(f"\nç™»å½•å‡ºé”™: {str(e)}")
                raise e
    
    client.start = custom_start
    
    print("æ­£åœ¨è¿æ¥åˆ° Telegram...")
    await client.start()
    print("å·²æˆåŠŸè¿æ¥ï¼")

    # è¯¢é—®æ˜¯å¦å¼€å§‹å¯¼å‡º
    while True:
        choice = input("\næ˜¯å¦å¼€å§‹å¯¼å‡ºæ•°æ®ï¼Ÿ(y/n): ").strip().lower()
        if choice == 'y':
            break
        elif choice == 'n':
            print("å·²å–æ¶ˆå¯¼å‡ºï¼Œç¨‹åºé€€å‡º")
            await client.disconnect()
            return
        else:
            print("æ— æ•ˆçš„è¾“å…¥ï¼Œè¯·è¾“å…¥ y æˆ– n")

    print("\nå¼€å§‹å¯¼å‡ºæ•°æ®...")
    export_time = datetime.now().strftime('%Y%m%d_%H%M%S')
    export_dir = f'tg_export_{export_time}'
    os.makedirs(export_dir, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # è·å–æ‰€æœ‰å¯¹è¯
    dialogs = await client.get_dialogs()
    
    try:
        if export_contacts:
            if stop_check and stop_check(): return
            print("\n[ä»»åŠ¡] æ­£åœ¨å¯¼å‡ºè”ç³»äººä¿¡æ¯")
            # è·å–æ‰€æœ‰å¯¹è¯ä¸­çš„ç”¨æˆ·
            users = []
            for dialog in dialogs:
                if isinstance(dialog.entity, User) and not dialog.entity.bot and not dialog.entity.deleted:
                    users.append(dialog.entity)
            
            print(f"[è¿›åº¦] å…±å‘ç° {len(users)} ä¸ªè”ç³»äºº")
            ws_contacts = wb.create_sheet("è”ç³»äºº")
            ws_contacts.append(['ç”¨æˆ·ID', 'ç”¨æˆ·å', 'æ‰‹æœºå·ç ', 'å§“å', 'çŠ¶æ€'])
            
            for i, user in enumerate(users, 1):
                if stop_check and stop_check(): return
                status = "æ­£å¸¸"
                if user.deleted:
                    status = "å·²æ³¨é”€"
                elif user.restricted:
                    # è·å–é™åˆ¶åŸå› 
                    reasons = []
                    if hasattr(user, 'restriction_reason'):
                        for restriction in user.restriction_reason:
                            reasons.append(restriction.text)
                    status = f"å—é™åˆ¶ ({', '.join(reasons)})" if reasons else "å—é™åˆ¶"
                ws_contacts.append([
                    str(user.id),
                    f"https://t.me/{user.username}" if user.username else '',
                    user.phone if hasattr(user, 'phone') else '',
                    f"{user.first_name or ''} {user.last_name or ''}".strip(),
                    status
                ])
                print(f"[è¿›åº¦] å·²å¤„ç† {i} ä¸ªè”ç³»äºº")
            
            adjust_column_width(ws_contacts)

        if export_groups:
            if stop_check and stop_check(): return
            print("\n[ä»»åŠ¡] æ­£åœ¨å¯¼å‡ºç¾¤ç»„ä¿¡æ¯")
            groups = [d for d in dialogs if d.is_group]
            print(f"[è¿›åº¦] å…±å‘ç° {len(groups)} ä¸ªç¾¤ç»„")
            ws_groups = wb.create_sheet("ç¾¤ç»„")
            ws_groups.append(['ç¾¤ç»„ID', 'ç¾¤ç»„åç§°', 'é‚€è¯·é“¾æ¥', 'çŠ¶æ€'])
            
            for i, group in enumerate(groups, 1):
                if stop_check and stop_check(): return
                entity = group.entity
                if isinstance(entity, (Chat, Channel)) and not getattr(entity, 'broadcast', False):
                    try:
                        status = "æ­£å¸¸"
                        if getattr(entity, 'left', False):
                            status = "å·²é€€å‡º"
                        elif getattr(entity, 'kicked', False):
                            status = "å·²è¢«è¸¢å‡º"
                        elif getattr(entity, 'restricted', False):
                            # è·å–é™åˆ¶åŸå› 
                            reasons = []
                            if hasattr(entity, 'restriction_reason'):
                                for restriction in entity.restriction_reason:
                                    reasons.append(restriction.text)
                            status = f"å—é™åˆ¶ ({', '.join(reasons)})" if reasons else "å—é™åˆ¶"
                        invite_link = await get_chat_invite_link(client, entity)
                        ws_groups.append([
                            str(entity.id),
                            entity.title,
                            invite_link,
                            status
                        ])
                    except Exception as e:
                        print(f"å¤„ç†ç¾¤ç»„æ—¶å‡ºé”™: {e}")
                print(f"[è¿›åº¦] å·²å¤„ç† {i} ä¸ªç¾¤ç»„")
            adjust_column_width(ws_groups)

        if export_channels:
            if stop_check and stop_check(): return
            print("\n[ä»»åŠ¡] æ­£åœ¨å¯¼å‡ºé¢‘é“ä¿¡æ¯")
            channels = [d for d in dialogs if isinstance(d.entity, Channel) and getattr(d.entity, 'broadcast', False)]
            print(f"[è¿›åº¦] å…±å‘ç° {len(channels)} ä¸ªé¢‘é“")
            ws_channels = wb.create_sheet("é¢‘é“")
            ws_channels.append(['é¢‘é“ID', 'é¢‘é“åç§°', 'é¢‘é“é“¾æ¥', 'çŠ¶æ€'])
            
            for i, dialog in enumerate(channels, 1):
                if stop_check and stop_check(): return
                entity = dialog.entity
                try:
                    status = "æ­£å¸¸"
                    if getattr(entity, 'left', False):
                        status = "å·²é€€å‡º"
                    elif getattr(entity, 'kicked', False):
                        status = "å·²è¢«è¸¢å‡º"
                    elif getattr(entity, 'restricted', False):
                        # è·å–é™åˆ¶åŸå› 
                        reasons = []
                        if hasattr(entity, 'restriction_reason'):
                            for restriction in entity.restriction_reason:
                                reasons.append(restriction.text)
                        status = f"å—é™åˆ¶ ({', '.join(reasons)})" if reasons else "å—é™åˆ¶"
                    channel_link = await get_chat_invite_link(client, entity)
                    ws_channels.append([
                        str(entity.id),
                        entity.title,
                        channel_link,
                        status
                    ])
                except Exception as e:
                    print(f"å¤„ç†é¢‘é“æ—¶å‡ºé”™: {e}")
                print(f"[è¿›åº¦] å·²å¤„ç† {i} ä¸ªé¢‘é“")
            adjust_column_width(ws_channels)

        if export_bots:
            if stop_check and stop_check(): return
            print("\n[ä»»åŠ¡] æ­£åœ¨å¯¼å‡ºæœºå™¨äººä¿¡æ¯")
            bots = [d for d in dialogs if isinstance(d.entity, User) and getattr(d.entity, 'bot', False)]
            print(f"[è¿›åº¦] å…±å‘ç° {len(bots)} ä¸ªæœºå™¨äºº")
            ws_bots = wb.create_sheet("æœºå™¨äºº")
            ws_bots.append(['æœºå™¨äººID', 'ç”¨æˆ·å', 'æœºå™¨äººåç§°', 'çŠ¶æ€'])
            
            for i, dialog in enumerate(bots, 1):
                if stop_check and stop_check(): return
                entity = dialog.entity
                try:
                    status = "æ­£å¸¸"
                    if getattr(entity, 'restricted', False):
                        status = "å·²é™åˆ¶"
                    ws_bots.append([
                        str(entity.id),
                        f"https://t.me/{entity.username}" if entity.username else '',
                        f"{entity.first_name or ''} {entity.last_name or ''}".strip(),
                        status
                    ])
                except Exception as e:
                    print(f"å¤„ç†æœºå™¨äººæ—¶å‡ºé”™: {e}")
                print(f"[è¿›åº¦] å·²å¤„ç† {i} ä¸ªæœºå™¨äºº")
            adjust_column_width(ws_bots)

        # ä¿å­˜Excelæ–‡ä»¶
        excel_file = os.path.join(export_dir, 'telegram_export.xlsx')
        wb.save(excel_file)
        print(f"\n[å®Œæˆ] å¯¼å‡ºå®Œæˆï¼æ–‡ä»¶ä¿å­˜åœ¨ {export_dir} ç›®å½•ä¸‹")

    finally:
        await client.disconnect()

if __name__ == '__main__':
    import asyncio
    asyncio.run(export_data()) 
